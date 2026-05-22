import logging
import os
import re
import shutil
import sys
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, cast

import openpyxl
import pandas as pd
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("F.Int.DataManager")


def split_complex_object(obj_text: str) -> List[str]:
    """Розбиває складні назви локацій на ізольовані атомарні об'єкти для фільтрації."""
    if not obj_text:
        return []
    obj_text = str(obj_text).strip()
    raw_parts = re.split(r";", obj_text)
    final_list = []
    for part in raw_parts:
        part = part.strip()
        if not part:
            continue
        match = re.search(r"^(.*?)\s*\(([^)]+)\)", part)
        if match:
            base_name = match.group(1).strip()
            inside_brackets = match.group(2)
            floors = re.findall(r"\d+", inside_brackets)
            if floors:
                for floor in floors:
                    final_list.append(f"{base_name} {floor} поверх")
            else:
                final_list.append(part)
        else:
            final_list.append(part)
    return list(dict.fromkeys(final_list))


class DataManager:
    def __init__(self, db_path: Optional[Path] = None):
        """Ініціалізація менеджера плоского атомарного реєстру майна."""
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            base_dir = Path(sys.executable).parent
        else:
            base_dir = Path(os.path.abspath(sys.argv[0])).parent

        self.info_dir = base_dir / "info"
        self.db_path = (
            db_path
            if db_path is not None
            else self.info_dir / "Structured_Asset_Base.xlsx"
        )
        self.cache_path = self.info_dir / "Structured_Asset_Base.pkl"

        self._df: pd.DataFrame = pd.DataFrame()
        self._file_not_found: bool = False

        self._ensure_infrastructure()
        self._load_data()

    def _ensure_infrastructure(self):
        try:
            if not self.info_dir.exists():
                self.info_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.error(f"Не вдалося ініціалізувати інфраструктуру папок: {e}")

    def _load_data(self):
        """Зчитування чистого плоского списку майна без деструктивних групувань Pandas."""
        if self.db_path.exists():
            try:
                df_excel = pd.read_excel(self.db_path, engine="openpyxl")
                if (
                    df_excel is not None
                    and isinstance(df_excel, pd.DataFrame)
                    and not df_excel.empty
                ):
                    self._df = df_excel
                    self._file_not_found = False
                    if "UUID" not in self._df.columns:
                        self._df.insert(
                            0, "UUID", [str(uuid.uuid4()) for _ in range(len(self._df))]
                        )
                        self.save_final_excel()
                    else:
                        self._df["UUID"] = self._df["UUID"].astype(str)
                    return
            except Exception as e:
                logger.error(f"Помилка завантаження Excel: {e}")
        self._file_not_found = True
        self._df = pd.DataFrame()

    def save_final_excel(self):
        """Синхронний атомарний запис реєстру на диск із застосуванням стилів openpyxl."""
        if self._df is None or self._df.empty:
            return
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        temp_excel = self.db_path.with_name(self.db_path.stem + "_tmp.xlsx")
        try:
            df_to_save = self._df.copy()
            if "Тип" in df_to_save.columns and "Тип майна" in df_to_save.columns:
                df_to_save = df_to_save.drop(columns=["Тип"])

            df_to_save.to_excel(temp_excel, index=False, engine="openpyxl")
            wb = openpyxl.load_workbook(temp_excel)
            ws = cast(Worksheet, wb.active)

            base_font = Font(name="Times New Roman", size=12)
            header_font = Font(
                name="Times New Roman", size=12, bold=True, color="FFFFFFFF"
            )
            header_fill = PatternFill(
                start_color="FF3B82F6", end_color="FF3B82F6", fill_type="solid"
            )
            alignment = Alignment(vertical="center", wrap_text=False)

            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                ws.row_dimensions[row_idx].height = 20
                for cell in row:
                    c = cast(Cell, cell)
                    c.alignment = alignment
                    if row_idx == 1:
                        c.font = header_font
                        c.fill = header_fill
                    else:
                        c.font = base_font

            for col in ws.columns:
                max_length = 0
                first_cell = cast(Cell, col[0])
                col_letter = get_column_letter(first_cell.column)
                for cell in col:
                    c = cast(Cell, cell)
                    if c.value:
                        max_length = max(max_length, len(str(c.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 55)

            if ws.dimensions:
                ws.auto_filter.ref = ws.dimensions
            wb.save(temp_excel)
            wb.close()
            os.replace(temp_excel, self.db_path)
            self._file_not_found = False
        except Exception as e:
            if os.path.exists(temp_excel):
                try:
                    os.remove(temp_excel)
                except Exception:
                    pass
            raise e

    def import_and_replace_db(self, source_file_path: str) -> Dict[str, Any]:
        try:
            src = Path(source_file_path)
            if not src.exists():
                return {"success": False, "error": "Обраний файл фізично не існує."}
            self.info_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, self.db_path)
            self._load_data()
            if self._file_not_found or self._df.empty:
                return {"success": False, "error": "Файл порожній або пошкоджений."}
            self.save_final_excel()
            return {"success": True, "rows_imported": len(self._df)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_aggregated_assets(self) -> List[Dict[str, Any]]:
        """Вивантажує повний плоский реєстр ТМЦ з UUID-деталізацією для фронтенду."""
        if self._file_not_found or self._df is None or self._df.empty:
            return [{"__SYSTEM_STATUS__": "FILE_NOT_FOUND"}]
        df_clean = self._df.fillna("")
        if "Кількість (факт)" in df_clean.columns:
            df_clean["Кількість (факт)"] = pd.to_numeric(
                df_clean["Кількість (факт)"], errors="coerce"
            ).fillna(0)
        raw_data = df_clean.to_dict(orient="records")
        result = []
        for row in raw_data:
            clean_row = {
                str(k): v for k, v in row.items() if not str(k).startswith("Unnamed")
            }
            clean_row.pop("Перекидка на Четверова", None)
            clean_row.pop("Перекидка на Іванова", None)
            if "Тип майна" in clean_row:
                clean_row["Тип"] = str(clean_row["Тип майна"]).strip()
            else:
                clean_row["Тип"] = ""
            obj_value = clean_row.get("Об'єкт", "")
            objects_list = split_complex_object(str(obj_value))
            clean_row["Об'єкт_список"] = objects_list
            result.append(clean_row)
        return result

    def add_asset(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        if self._df is None or self._df.empty:
            default_columns = [
                "UUID",
                "Підрозділ",
                "МВО (Прізвище)",
                "Об'єкт",
                "Тип майна",
                "Найменування",
                "Кількість (факт)",
                "Одиниця виміру",
            ]
            self._df = pd.DataFrame(columns=default_columns)

        new_uuid = str(uuid.uuid4())
        new_row: Dict[str, Any] = {"UUID": new_uuid}
        if "Тип" in payload:
            payload["Тип майна"] = payload["Тип"]

        for col in self._df.columns:
            if col == "UUID" or col == "Тип":
                continue
            if col in payload:
                if pd.api.types.is_numeric_dtype(self._df[col]):
                    try:
                        new_row[col] = float(payload[col]) if payload[col] else 0.0
                    except ValueError:
                        new_row[col] = 0.0
                else:
                    new_row[col] = str(payload[col])
            else:
                new_row[col] = (
                    0.0 if pd.api.types.is_numeric_dtype(self._df[col]) else ""
                )

        self._df = pd.concat([self._df, pd.DataFrame([new_row])], ignore_index=True)

        # ВИПРАВЛЕНО: Алфавітне сортування всього реєстру за колонкою 'Тип майна'
        try:
            if "Тип майна" in self._df.columns:
                self._df = self._df.sort_values(
                    by="Тип майна", ascending=True, na_position="last"
                ).reset_index(drop=True)
            self.save_final_excel()
            return {"success": True, "uuid": new_uuid}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def edit_asset(
        self, asset_uuid: str, payload: Dict[str, Any], mode: str
    ) -> Dict[str, Any]:
        if self._df is None or self._df.empty:
            return {"success": False, "error": "Реєстр порожній."}
        idx = self._df.index[self._df["UUID"].astype(str) == str(asset_uuid)]
        if len(idx) == 0:
            return {"success": False, "error": f"Актив не знайдено: {asset_uuid}"}

        row_idx = idx[0]
        if "Тип" in payload:
            payload["Тип майна"] = payload["Тип"]

        for key, value in payload.items():
            if key in ["UUID", "Об'єкт_список", "Тип"]:
                continue
            if key in self._df.columns:
                self._df[key] = self._df[key].astype(object)
                if pd.api.types.is_numeric_dtype(self._df[key]):
                    try:
                        self._df.at[row_idx, key] = float(value) if value else 0.0
                    except ValueError:
                        self._df.at[row_idx, key] = 0.0
                else:
                    self._df.at[row_idx, key] = str(value) if value is not None else ""
        try:
            self.save_final_excel()
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def delete_asset(self, asset_uuid: str) -> Dict[str, Any]:
        if self._df is None or self._df.empty:
            return {"success": False, "error": "Реєстр порожній."}
        idx = self._df.index[self._df["UUID"].astype(str) == str(asset_uuid)]
        if len(idx) == 0:
            return {"success": False, "error": "Актив не знайдено."}

        self._df = self._df.drop(index=idx).reset_index(drop=True)
        try:
            self.save_final_excel()
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def add_custom_column(self, name: str) -> Dict[str, Any]:
        name = name.strip()
        if not name:
            return {"success": False, "error": "Назва пожня."}
        if name in self._df.columns or name == "Тип":
            return {"success": False, "error": f"Пункт '{name}' вже існує."}
        self._df[name] = ""
        self.save_final_excel()
        return {"success": True}

    def delete_custom_column(self, name: str) -> Dict[str, Any]:
        name = name.strip()
        protected = ["UUID", "Тип майна", "МВО (Прізвище)", "Об'єкт", "Тип"]
        if name in protected:
            return {"success": False, "error": f"Заборонено видаляти пункт '{name}'."}
        if name not in self._df.columns:
            return {"success": False, "error": f"Пункт '{name}' не знайдено."}
        self._df = self._df.drop(columns=[name])
        self.save_final_excel()
        return {"success": True}
