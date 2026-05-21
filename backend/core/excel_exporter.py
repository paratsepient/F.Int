import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.cell.cell import Cell  # ← один імпорт вгорі
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("F.Int.ExcelExporter")


def _set_cell(ws, row: int, col: int, value=None, **styles) -> None:
    """Встановлює значення та стилі тільки для справжніх Cell (не MergedCell)."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, Cell):  # MergedCell — пропускаємо
        return
    if value is not None:
        cell.value = value
    for attr, val in styles.items():
        setattr(cell, attr, val)


class ExcelExporter:
    def __init__(self, archive_dir: Path, data_manager: Any):
        self.archive_dir = archive_dir
        self.data_manager = data_manager
        self.archive_dir.mkdir(exist_ok=True)

    def generate_document(
        self, uuids: List[str], payload: Dict[str, Any], mode: str
    ) -> Dict[str, Any]:
        try:
            all_assets = self.data_manager.get_aggregated_assets()
            selected_assets = [a for a in all_assets if str(a.get("UUID")) in uuids]

            if not selected_assets:
                return {"success": False, "error": "Не знайдено майна для експорту."}

            title = payload.get(
                "title", f"Документ_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )
            filename = f"{title.replace(' ', '_')}.xlsx"
            file_path = self.archive_dir / filename

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                return {"success": False, "error": "Не вдалося створити аркуш Excel."}
            ws.title = "Акт"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Заголовки
            headers = ["Тип", "Найменування", "Інв. №", "Об'єкт", "МВО"]
            for col_num, header in enumerate(headers, 1):
                _set_cell(
                    ws,
                    1,
                    col_num,
                    value=str(header),
                    font=header_font,
                    fill=header_fill,
                    alignment=Alignment(horizontal="center"),
                )

            # Дані
            for row_idx, asset in enumerate(selected_assets, 2):
                values = [
                    str(asset.get("Тип", "") or ""),
                    str(asset.get("Найменування", "") or ""),
                    str(asset.get("Інв. / Номенкл. №", "") or ""),
                    str(asset.get("Об'єкт", "") or ""),
                    str(asset.get("МВО (Прізвище)", "") or ""),
                ]
                for col_idx, val in enumerate(values, 1):
                    _set_cell(ws, row_idx, col_idx, value=val, border=thin_border)

            # Автоширина
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

            wb.save(file_path)
            logger.info(f"Документ згенеровано: {file_path}")
            return {"success": True, "filename": filename}

        except Exception as e:
            logger.error(f"Помилка генерації Excel: {e}")
            return {"success": False, "error": str(e)}
